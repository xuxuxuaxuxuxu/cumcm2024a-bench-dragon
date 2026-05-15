#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
2024年高教社杯全国大学生数学建模竞赛 A题 "板凳龙"闹元宵
完整求解代码 (修正版 v2.4 - 优化版)

主要修正:
  v2.0 (v1→v2):
    1. 链式约束方向修正: 从龙头到龙尾的链方向与龙头运动方向相反
    2. 问题4: 使用路径弧长参数化方法处理S曲线上的链约束
    3. 问题4: 盘出螺线段也使用链约束求解而非线性插值
    4. 运算符优先级修正: r = sqrt(x² + y²)

  v2.1:
    5. 弦长约束修正: 用 |Pₖ₊₁ - Pₖ| = Lₖ (弦长) 替代弧长偏移假设
       - 新增 get_unified_path_point(): 统一路径弧长→坐标映射
       - 新增 solve_next_handle_on_path(): 数值二分法求解弦长约束
       - 新增 compute_chain_on_unified_path(): 统一路径上的完整链求解
    6. 碰撞检测范围扩展: 从邻近15节扩展到全体非相邻板凳
    7. 盘入阶段(t<0)改用DragonSimulation确保正确弦长约束
    8. 清理: 移除未使用的 compute_chain_on_path 和
       compute_chain_on_composite_path 函数

  v2.2:
    9.  碰撞检测修正: _point_in_convex_polygon 现在自动适应CW/CCW顶点
        顺序(原代码仅支持CCW, 但矩形顶点按CW构建, 导致碰撞漏检)。
    10. 碰撞检测增强: 新增边交叉检测(_segments_intersect), 捕获无顶点
        包含的十字交叉碰撞情形。

  v2.3:
    11. 盘出螺线定义修正: 题目要求盘出螺线是盘入螺线的"中心对称像",
        即 E(θ) = -P(θ)。原代码使用 spiral_point(theta_entry+π) 和
        spiral_point(theta, b), 导致盘出螺线整体偏离正确位置约
        b·π ≈ 0.85m/圈。现修正为 P_exit = -P_entry,
        盘出螺线点 = -spiral_point(theta, b)。
        (相关函数: CompositePath.__init__, get_unified_path_point)
    12. S曲线C¹连续性再修正: v2.2第(9)项修正方向错误。
        正确配置应为 Arc1(CW反转) + Arc2(CCW正转):
        - Arc1 CW: 在P_entry处切线 = (sinθₑ, -cosθₑ) → 匹配盘入螺线 ✓
        - Arc1 CCW: 在P_entry处切线 = (-sinθₑ, cosθₑ) → 相反 ✗
        - Arc2 CCW: 在出口P_exit处切线匹配盘出螺线(负切向) ✓
        (相关函数: CompositePath._compute_arc_angles)

  v2.4 (性能优化):
    13. Hot-start链求解: compute_chain_positions 接收 prev_thetas
        参数, 用上一时间步的theta作为当前步的初始猜测(fsolve),
        brentq作为回退。相邻时间步龙队位移很小, fsolve 通常
        1-2步收敛, 较brentq快约25倍。
        (相关函数: forward_simulation, simulate_at_time,
         compute_chain_positions, solve_problem2, solve_problem4)
    14. Theta直接求解盘出螺线: compute_chain_on_unified_path 在
        盘入/盘出螺线段跟踪theta, 直接使用 solve_next_handle_theta
        求解弦长约束, 避免嵌套brentq(路径求解器每步调用的
        get_unified_path_point在螺线段内部还有一层brentq)。
        在盘出螺线段(约52节/时间步)和盘入螺线段(约164节/时间步)
        生效, 每步节省约216次内层brentq。
        同时传入解析 theta_guess 触发 fsolve 快速路径,
        进一步避免brentq调用。
    15. solve_theta_from_arc_length 紧区间: 使用ds/dθ≈b·θ估计
        Δθ, 将搜索区间从[θ₀, θ₀+1000]缩小到[θ₀, θ₀+Δθ·1.5],
        减少brentq迭代次数。
    16. solve_next_handle_theta 优化: 预计算P_current到循环外,
        避免目标函数中重复计算; 使用弧长近似缩小初始brentq区间。
    17. 速度计算全覆盖量化: compute_velocities 用numpy切片操作
        替代Python循环, 对含n_times个样本的数组一次完成。
    18. 碰撞检测AABB预检: _polygons_intersect 先做外接矩形重叠
        快速拒绝, 再执行O(N²)精细检测。
    19. Excel输出覆盖量化: 用 np.round(数组,6) 替代逐元素
        Python round() 调用。
    20. 清理: 移除未使用的 scipy.integrate.quad 导入;
        修复 main() 中版本号字符串;
        修复 solve_problem4 中 path.theta_exit 属性不存在错误。

运行环境: Python 3.8+, 依赖: numpy, scipy, openpyxl, matplotlib
运行命令: python bench_dragon_simulation.py
"""

import numpy as np
from scipy.optimize import brentq, fsolve
import openpyxl
import matplotlib.pyplot as plt
import warnings
import os
import time

warnings.filterwarnings('ignore')

# ============================================================
# 第0部分: 基础参数与几何函数
# ============================================================

HEAD_LENGTH = 3.41
BODY_LENGTH = 2.20
WIDTH = 0.30
HOLE_OFFSET = 0.275

HEAD_HOLE_SPACING = HEAD_LENGTH - 2 * HOLE_OFFSET  # 2.86 m
BODY_HOLE_SPACING = BODY_LENGTH - 2 * HOLE_OFFSET  # 1.65 m

N_SEGMENTS = 223
N_HANDLES = N_SEGMENTS + 1  # 224

# SPACINGS[i] = 把手i到把手i+1的距离
# i=0: 2.86m, i=1..222: 1.65m
SPACINGS = np.array([HEAD_HOLE_SPACING] + [BODY_HOLE_SPACING] * (N_SEGMENTS - 1))

# 累计弧长 (从龙头前把手到第i个把手的路径总长)
CUMULATIVE_LENGTH = np.zeros(N_HANDLES)
for i in range(1, N_HANDLES):
    CUMULATIVE_LENGTH[i] = CUMULATIVE_LENGTH[i - 1] + SPACINGS[i - 1]
# CUMULATIVE_LENGTH[0] = 0
# CUMULATIVE_LENGTH[1] = 2.86
# CUMULATIVE_LENGTH[2] = 2.86 + 1.65 = 4.51
# ...
# CUMULATIVE_LENGTH[223] ≈ 2.86 + 222*1.65 = 369.16 m

DRAGON_TOTAL_LENGTH = CUMULATIVE_LENGTH[-1]  # 龙队全长 (从龙头前到龙尾后)


def get_b(pitch):
    """螺线参数 b = pitch / (2π)"""
    return pitch / (2 * np.pi)


# ============================================================
# 螺线几何函数
# ============================================================

def spiral_point(theta, b):
    """阿基米德螺线上的点坐标 (标量theta)"""
    r = b * theta
    return np.array([r * np.cos(theta), r * np.sin(theta)])


def spiral_points(thetas, b):
    """向量化: 批量计算多个theta对应的螺线坐标 (N,) → (N, 2)"""
    r = b * thetas
    return np.column_stack([r * np.cos(thetas), r * np.sin(thetas)])


def spiral_arc_length_func(theta, b):
    """弧长解析函数: s(θ) = (b/2)[θ√(θ²+1) + arcsinh(θ)]"""
    return (b / 2) * (theta * np.sqrt(theta ** 2 + 1) + np.arcsinh(theta))


def spiral_arc_length(theta_start, theta_end, b):
    """从 theta_start 到 theta_end 的弧长"""
    return spiral_arc_length_func(theta_end, b) - spiral_arc_length_func(theta_start, b)


def solve_theta_from_arc_length(theta_initial, target_arc_length, b, direction=-1):
    """
    从 theta_initial 出发，沿 direction 前进 target_arc_length 弧长，
    求解到达的 theta 值
    direction: -1 顺时针盘入 (theta减小), +1 逆时针盘出 (theta增大)
    """
    s_initial = spiral_arc_length_func(theta_initial, b)
    s_target = s_initial + direction * target_arc_length

    if s_target < 0:
        return None

    # 使用弧长近似计算更紧的初始区间
    # ds/dθ = b*sqrt(θ²+1) ≈ b*θ (大θ时)
    # Δθ ≈ target_arc_length / (b * sqrt(θ² + 1))
    deriv = b * np.sqrt(theta_initial ** 2 + 1)
    delta_est = target_arc_length / max(deriv, b)

    if direction < 0:
        theta_low = max(1e-12, theta_initial - delta_est * 1.5)
        theta_high = theta_initial
        # 缩小下界确保有根
        if theta_low < 1e-12:
            theta_low = 1e-12
    else:
        theta_low = theta_initial
        theta_high = theta_initial + max(0.5, delta_est * 1.5)

    def f(theta):
        return spiral_arc_length_func(theta, b) - s_target

    # 检查边界
    f_low, f_high = f(theta_low), f(theta_high)
    if f_low * f_high > 0:
        # 扩界
        for _ in range(100):
            if direction < 0:
                theta_low = max(1e-14, theta_low * 0.5)
            else:
                theta_high += max(0.5, delta_est)
            f_low, f_high = f(theta_low), f(theta_high)
            if f_low * f_high <= 0:
                break
        else:
            return None  # 无法找到有根区间

    try:
        return brentq(f, theta_low, theta_high, xtol=1e-12)
    except (ValueError, RuntimeError):
        try:
            return brentq(f, theta_low, theta_high, xtol=1e-10)
        except:
            return None


# --- 链式约束求解 (螺线上) ---

def solve_next_handle_theta(theta_current, target_distance, b, direction=1, theta_guess=None):
    """
    已知当前把手在 theta_current，求下一个把手使距离 = target_distance
    direction: +1 (theta增大, 向外), -1 (theta减小, 向内)

    优化: 预计算 P_current 避免在目标函数中重复计算;
         使用弧长近似缩小初始区间以加速 brentq 收敛.
    """
    # 预计算当前把手位置 (避免f中反复计算)
    P_current = spiral_point(theta_current, b)

    def f(theta_next):
        P_next = spiral_point(theta_next, b)
        return np.linalg.norm(P_next - P_current) - target_distance

    # 弧长近似: ds/dθ ≈ b * sqrt(θ² + 1)
    deriv = b * np.sqrt(theta_current ** 2 + 1)
    delta_est = target_distance / max(deriv, b)

    if direction > 0:  # theta_next > theta_current (向外)
        theta_low = theta_current
        theta_high = theta_current + max(1.0, delta_est * 2.0)

        # 尝试初始猜测 (hot-start / 解析近似)
        if theta_guess is not None and theta_guess > theta_current:
            sol = fsolve(f, theta_guess, maxfev=500, xtol=1e-10)
            if abs(f(sol[0])) < 1e-8:
                return sol[0]

        # 二分法
        n_tries = 0
        f_low, f_high = f(theta_low), f(theta_high)
        while f_low * f_high > 0 and n_tries < 100:
            theta_high += max(1.0, delta_est)
            f_high = f(theta_high)
            n_tries += 1

        if f_low * f_high <= 0:
            try:
                return brentq(f, theta_low, theta_high, xtol=1e-12)
            except:
                pass

        return None

    else:  # direction < 0, theta_next < theta_current (向内)
        theta_low = max(1e-12, theta_current - max(1.0, delta_est * 2.0))
        theta_high = theta_current

        if theta_guess is not None and theta_low < theta_guess < theta_high:
            sol = fsolve(f, theta_guess, maxfev=500, xtol=1e-10)
            if abs(f(sol[0])) < 1e-8:
                return sol[0]

        # 二分法 (初始区间已收紧)
        n_tries = 0
        f_low, f_high = f(theta_low), f(theta_high)
        while f_low * f_high > 0 and n_tries < 100:
            theta_low = max(1e-14, theta_low * 0.5)
            f_low = f(theta_low)
            n_tries += 1

        if f_low * f_high <= 0:
            try:
                return brentq(f, theta_low, theta_high, xtol=1e-12)
            except:
                pass

        return None


# ============================================================
# 第1部分: 模拟引擎 (修正版)
# ============================================================

class DragonSimulation:
    """
    舞龙队模拟引擎

    物理约定:
    - 龙头 (handle 0) 在前引导，龙尾 (handle 223) 在后跟随
    - 盘入时: 龙头向内 (θ减小), 沿链向后则 θ增大 (向外)
    - 盘出时: 龙头向外 (θ增大), 沿链向后则 θ减小 (向内)
    - 因此链方向 = -龙头运动方向
    """
    def __init__(self, b, initial_theta_head, head_direction=-1):
        """
        b: 螺线参数
        initial_theta_head: 龙头前把手初始角度
        head_direction: 龙头运动方向 (-1盘入, +1盘出)
        """
        self.b = b
        self.initial_theta = initial_theta_head
        self.head_direction = head_direction
        self.chain_direction = -head_direction  # 链方向与龙头方向相反!
        self.thetas = None
        self.positions = None
        self.velocities = None
        self.speeds = None
        self.times = None

    def compute_chain_positions(self, theta_head, prev_thetas=None):
        """
        给定龙头前把手角度 theta_head，计算所有把手角度
        从龙头开始沿链向后: chain_direction决定theta增减方向
        prev_thetas: 上一时间步的theta数组，用于hot-start加速
        """
        n = N_HANDLES
        thetas = np.zeros(n)
        thetas[0] = theta_head

        for i in range(n - 1):
            target_dist = SPACINGS[i]

            # 使用上一时间步的值作为初始猜测 (hot-start)
            if prev_thetas is not None and i + 1 < len(prev_thetas):
                theta_guess = prev_thetas[i + 1]
            else:
                # 解析近似作为初始猜测
                if thetas[i] > 1e-6:
                    theta_guess = (thetas[i] +
                                   self.chain_direction * target_dist /
                                   (self.b * np.sqrt(thetas[i]**2 + 1)))
                else:
                    theta_guess = thetas[i] + self.chain_direction * target_dist / self.b

            result = solve_next_handle_theta(
                thetas[i], target_dist, self.b,
                direction=self.chain_direction, theta_guess=theta_guess
            )

            if result is None:
                return None
            thetas[i + 1] = result

        return thetas

    def simulate_at_time(self, t, head_speed=1.0, prev_thetas=None):
        """
        计算 t 时刻龙队的状态
        prev_thetas: 上一时间步的theta数组，用于hot-start
        返回: (thetas, positions)
        """
        arc_length = head_speed * t

        theta_head = solve_theta_from_arc_length(
            self.initial_theta, arc_length, self.b, direction=self.head_direction
        )
        if theta_head is None:
            return None

        thetas = self.compute_chain_positions(theta_head, prev_thetas)
        if thetas is None:
            return None

        positions = spiral_points(thetas, self.b)
        return thetas, positions

    def forward_simulation(self, times, head_speed=1.0, verbose=True):
        """
        前向模拟，计算多个时间步
        使用 hot-start: 前一步的theta作为下一步的初始猜测，加速收敛
        返回: (thetas_all, positions_all, times_out)
        """
        n_times = len(times)
        thetas_all = np.zeros((n_times, N_HANDLES))
        positions_all = np.zeros((n_times, N_HANDLES, 2))

        prev_thetas = None
        for k, t in enumerate(times):
            # 使用 hot-start: 前一步的theta传给compute_chain_positions
            result = self.simulate_at_time(t, head_speed, prev_thetas)

            if result is None:
                if verbose:
                    print(f"  [T={t:.1f}s] 模拟终止：链式约束求解失败")
                return thetas_all[:k], positions_all[:k], times[:k]

            thetas, positions = result
            prev_thetas = thetas  # 保存供下一步使用
            thetas_all[k] = thetas
            positions_all[k] = positions

            if verbose and k % 50 == 0:
                r = np.linalg.norm(positions[0])
                print(f"  [T={t:.1f}s] 龙头 theta={thetas[0]:.4f} rad, r={r:.4f} m")

        return thetas_all, positions_all, times

    @staticmethod
    def compute_velocities(positions_all, times):
        """
        计算所有把手的速度 (三点中心差分, 全覆盖量化)
        """
        n_times = len(times)
        speeds_all = np.zeros((n_times, N_HANDLES))

        if n_times < 2:
            return speeds_all

        # 前向差分: 第一个时间步
        dt0 = times[1] - times[0]
        speeds_all[0] = np.linalg.norm(positions_all[1] - positions_all[0], axis=1) / dt0

        # 中心差分: 中间时间步 (全覆盖量化, 无Python循环)
        if n_times > 2:
            dt_center = times[2:] - times[:-2]
            speeds_all[1:-1] = np.linalg.norm(
                positions_all[2:] - positions_all[:-2], axis=2
            ) / dt_center[:, np.newaxis]

        # 后向差分: 最后一个时间步
        dt_last = times[-1] - times[-2]
        speeds_all[-1] = np.linalg.norm(positions_all[-1] - positions_all[-2], axis=1) / dt_last

        return speeds_all


# ============================================================
# 第1.5部分: 组合路径 (S曲线 + 螺旋线) 链约束求解器
# 用于问题4和问题5
# ============================================================

class CompositePath:
    """
    组合路径: 盘入螺线 + S形调头曲线 + 盘出螺线
    所有段在连接点处 C¹ 连续 (位置和切向连续)
    提供弧长参数化: point_at_arc_length(s) → (x, y)
    """

    def __init__(self, pitch, R_turn=4.5):
        self.pitch = pitch
        self.b = get_b(pitch)
        self.R_turn = R_turn

        # 盘入螺线与调头空间边界的交点
        self.theta_entry = R_turn / self.b
        self.P_entry = spiral_point(self.theta_entry, self.b)

        # 盘出起点: 中心对称 (非同一螺线上的 θ+π!)
        # 出口螺线 E(θ) = -P(θ), 故 P_exit = -P_entry
        # (若使用 spiral_point(theta_entry+π) 将偏出 b*π ≈ 0.85m)
        self.P_exit = -self.P_entry

        # 入口/出口法向 (指向圆心)
        self.n_entry = -self.P_entry / np.linalg.norm(self.P_entry)
        self.n_exit = -self.P_exit / np.linalg.norm(self.P_exit)

        # 入口/出口切向 (出口螺线 E(θ) = -P(θ) 的切线)
        self.t_entry = self._spiral_tangent(self.theta_entry)
        self.t_exit = -self._spiral_tangent(self.theta_entry)

        # 求解S曲线半径 (给定 R1 = 2*R2)
        self._solve_s_curve()

        # 各段弧长
        self.arc1_len = abs(self.arc1_angle) * self.R1
        self.arc2_len = abs(self.arc2_angle) * self.R2
        self.s_total = self.arc1_len + self.arc2_len

    def _spiral_tangent(self, theta):
        """螺线在角度theta处的单位切向量 (指向θ增大的方向)"""
        dx = np.cos(theta) - theta * np.sin(theta)
        dy = np.sin(theta) + theta * np.cos(theta)
        t = np.array([dx, dy])
        return t / np.linalg.norm(t)

    def _angle_diff(self, a1, a2):
        """a2 - a1 在 (-π, π] 范围内的有向角差"""
        d = (a2 - a1) % (2 * np.pi)
        if d > np.pi:
            d -= 2 * np.pi
        return d

    def _solve_s_curve(self):
        """求解S形调头曲线的两个圆弧半径"""
        # 入口处法向指向圆心
        # 圆心 O1 = P_entry + R1 * n_entry = P_entry + 2*R2 * n_entry
        # 圆心 O2 = P_exit + R2 * n_exit
        # 相切条件: |O2 - O1| = R1 + R2 = 3*R2

        def eq(R2):
            O1 = self.P_entry + 2 * R2 * self.n_entry
            O2 = self.P_exit + R2 * self.n_exit
            return np.linalg.norm(O2 - O1) - 3 * R2

        # 二分法求R2
        R_low, R_high = 0.01, self.R_turn

        # 找到有根区间
        for _ in range(100):
            if eq(R_low) * eq(R_high) <= 0:
                break
            R_high *= 0.5

        if eq(R_low) * eq(R_high) > 0:
            # 尝试扩大范围
            R_high = self.R_turn
            for _ in range(100):
                R_high *= 1.5
                if eq(R_low) * eq(R_high) <= 0:
                    break

            if eq(R_low) * eq(R_high) > 0:
                print(f"  警告: S曲线半径求解器未找到有根区间, 使用默认值")
                self.R2 = self.R_turn * 0.3
                self.R1 = 2 * self.R2
                self._compute_arc_angles()
                return

        try:
            self.R2 = brentq(eq, R_low, R_high, xtol=1e-10)
        except:
            self.R2 = (R_low + R_high) / 2

        self.R1 = 2 * self.R2
        self._compute_arc_angles()

    def _compute_arc_angles(self):
        """计算圆弧起止角度和中间点"""
        self.O1 = self.P_entry + self.R1 * self.n_entry
        self.O2 = self.P_exit + self.R2 * self.n_exit

        # 切点 (在 O1-O2 连线上)
        vec_O1_O2 = self.O2 - self.O1
        dir_O1_O2 = vec_O1_O2 / np.linalg.norm(vec_O1_O2)
        self.P_mid = self.O1 + self.R1 * dir_O1_O2

        # --- Arc1 (入口→中点): 反转(CW)使入口切线匹配盘入螺线 ---
        # 盘入螺线在P_entry处切线 ≈ (sinθₑ, -cosθₑ) [θₑ较大时主导项]
        # CCW弧在P_entry处切线 = (-sinθₑ, cosθₑ) → 相反, 不匹配
        # CW弧在P_entry处切线 = (sinθₑ, -cosθₑ) → 匹配盘入螺线 ✓
        v_start = self.P_entry - self.O1
        self.angle_start1 = np.arctan2(v_start[1], v_start[0])
        v_mid1 = self.P_mid - self.O1
        self.angle_mid1 = np.arctan2(v_mid1[1], v_mid1[0])
        self.arc1_angle = self._angle_diff(self.angle_start1, self.angle_mid1)
        # Arc1反转(CW)
        if self.arc1_angle > 0:
            self.arc1_angle = -self.arc1_angle

        # --- Arc2 (中点→出口): 正转(CCW)使中点切线匹配Arc1 ---
        # Arc1(CW)在中点切线 = (-sinθₑ, cosθₑ)
        # Arc2(CCW)在中点切线 = (-sinθₑ, cosθₑ) → 匹配 ✓
        # Arc2(CCW)在出口切线 = (sinθₑ, -cosθₑ) → 匹配盘出螺线 ✓
        v_mid2 = self.P_mid - self.O2
        self.angle_start2 = np.arctan2(v_mid2[1], v_mid2[0])
        v_end = self.P_exit - self.O2
        self.angle_end2 = np.arctan2(v_end[1], v_end[0])
        self.arc2_angle = self._angle_diff(self.angle_start2, self.angle_end2)
        # Arc2保持正转(CCW), 已为正无需修正

    def point_on_curve(self, s):
        """
        给定从入口沿S曲线的弧长 s ∈ [0, s_total]，
        返回位置向量 (x, y)
        """
        if s <= self.arc1_len:
            frac = s / self.arc1_len if self.arc1_len > 0 else 0
            angle = self.angle_start1 + frac * self.arc1_angle
            return self.O1 + self.R1 * np.array([np.cos(angle), np.sin(angle)])
        else:
            s2 = s - self.arc1_len
            frac = s2 / self.arc2_len if self.arc2_len > 0 else 0
            angle = self.angle_start2 + frac * self.arc2_angle
            return self.O2 + self.R2 * np.array([np.cos(angle), np.sin(angle)])


# ============================================================
# 统一路径弧长坐标系 + 弦长约束求解器
# 用于问题4的S曲线和盘出螺线段
# ============================================================

def get_unified_path_point(s, path_obj):
    """
    统一路径弧长坐标 s → (x, y) 映射

    弧长坐标系:
        s = 0:  P_entry (盘入螺线与S曲线交点)
        s < 0:  盘入螺线上 (向外延伸)
        s ∈ (0, s_total): S曲线上
        s > s_total: 盘出螺线上 (向外延伸)

    约束: |P(s₂) - P(s₁)| < |s₂ - s₁| 对所有 s₁ ≠ s₂ 成立
          (因为路径是曲线的, 弦长 ≤ 弧长)
    """
    b = path_obj.b

    if s < 0:
        # 盘入螺线: 从 P_entry 向外走 |s| 弧长
        s_out = -s
        theta = solve_theta_from_arc_length(
            path_obj.theta_entry, s_out, b, direction=1
        )
        if theta is None:
            return None
        return spiral_point(theta, b)

    elif s <= path_obj.s_total:
        # S曲线上
        return path_obj.point_on_curve(s)

    else:
        # 盘出螺线: 从 P_exit 向外走 (s - s_total) 弧长
        # 出口螺线 E(θ) = -P(θ), 是入口螺线的中心对称像
        # 映射: 入口螺线上从 theta_entry 向外走 s_out 弧长 → θ → 取负
        s_out = s - path_obj.s_total
        theta = solve_theta_from_arc_length(
            path_obj.theta_entry, s_out, b, direction=1
        )
        if theta is None:
            return None
        # E(θ) = -P(θ), 中心对称
        return -spiral_point(theta, b)


def solve_next_handle_on_path(current_s, target_dist, path_obj,
                              direction=-1, max_iter=200):
    """
    在统一路径上求解弦长约束。

    给定 current_s 处的把手位置 P_current,
    求 s_next 使 |P(s_next) - P(current_s)| = target_dist.

    direction: -1 向后 (tail方向, s减小), +1 向前 (head方向, s增大)

    原理: 使用二分法求解一维非线性方程 |P(s) - P(current_s)| - target_dist = 0
    由于弦长单调地随 |s - current_s| 增加，二分法保证收敛。
    """
    def chord_length(s):
        """计算从 current_s 到 s 的弦长"""
        P_curr = get_unified_path_point(current_s, path_obj)
        P_next = get_unified_path_point(s, path_obj)
        if P_curr is None or P_next is None:
            return -1.0
        return np.linalg.norm(P_next - P_curr)

    P_current = get_unified_path_point(current_s, path_obj)
    if P_current is None:
        return None

    def f(s):
        P_s = get_unified_path_point(s, path_obj)
        if P_s is None:
            return -target_dist
        return np.linalg.norm(P_s - P_current) - target_dist

    if direction < 0:
        # 向后 (tail方向): 寻找 s_next < current_s
        s_high = current_s
        s_low = current_s - target_dist * 2

        # f(current_s) = -target_dist < 0
        # 需要 f(s_low) > 0 (弦长 > target_dist)
        for _ in range(max_iter):
            f_low = f(s_low)
            if f_low > 0:
                break
            s_low -= target_dist

        if f_low <= 0:
            return None  # 无法找到有根区间

        try:
            s_sol = brentq(f, s_low, s_high, xtol=1e-10)
            return s_sol
        except (ValueError, RuntimeError):
            return None

    else:
        # 向前: 寻找 s_next > current_s
        s_low = current_s
        s_high = current_s + target_dist * 2

        # f(current_s) = -target_dist < 0
        # 需要 f(s_high) > 0
        for _ in range(max_iter):
            f_high = f(s_high)
            if f_high > 0:
                break
            s_high += target_dist

        if f_high <= 0:
            return None

        try:
            s_sol = brentq(f, s_low, s_high, xtol=1e-10)
            return s_sol
        except (ValueError, RuntimeError):
            return None


def compute_chain_on_unified_path(head_arc_pos, path_obj):
    """
    在统一路径上使用弦长约束求解龙队位置。
    优化版: 在螺旋线段使用 theta 直接求解，避免嵌套 brentq。

    输入:
        head_arc_pos: 龙头前把手的弧长坐标
                      (>0 for S曲线/盘出, <0 for 盘入)
        path_obj: CompositePath 实例

    输出:
        positions: (N_HANDLES, 2) 或 None
    """
    b = path_obj.b
    s_total = path_obj.s_total
    theta_entry = path_obj.theta_entry
    EP = 1e-10

    positions = np.zeros((N_HANDLES, 2))

    # handle 0 (龙头前把手)
    P0 = get_unified_path_point(head_arc_pos, path_obj)
    if P0 is None:
        return None
    positions[0] = P0

    # 跟踪当前弧长坐标和theta (仅螺旋线段)
    s_curr = head_arc_pos
    theta_curr = None  # 仅当在螺旋线段时非空

    # 若头部在螺旋线段，计算初值theta
    if abs(s_curr) > s_total + EP:
        s_out = abs(s_curr) - s_total
        theta_curr = solve_theta_from_arc_length(
            theta_entry, s_out, b, direction=1
        )

    # 逐节求解
    for i in range(N_HANDLES - 1):
        target = SPACINGS[i]

        # --- 情形1: 在盘出螺旋线 (s > s_total) 上用 theta 直接求解 ---
        if s_curr > s_total + EP and theta_curr is not None:
            # 尾方向: theta递减; 解析初值触发fsolve快速路径
            guess_out = theta_curr - target / (b * np.sqrt(theta_curr ** 2 + 1))
            theta_next = solve_next_handle_theta(
                theta_curr, target, b, direction=-1,
                theta_guess=guess_out
            )
            if theta_next is None:
                return None
            # 检测是否即将进入S曲线
            if theta_next < theta_entry * 1.001:
                # 过渡到S曲线: 回退到路径求解器
                s_next = solve_next_handle_on_path(
                    s_curr, target, path_obj, direction=-1
                )
                if s_next is None:
                    return None
                positions[i + 1] = get_unified_path_point(s_next, path_obj)
                s_curr = s_next
                theta_curr = None
            else:
                # 正常盘出螺线段: theta直接→坐标
                arc = spiral_arc_length(theta_entry, theta_next, b)
                s_curr = s_total + arc
                positions[i + 1] = -spiral_point(theta_next, b)
                theta_curr = theta_next

        # --- 情形2: 在盘入螺旋线 (s < 0) 上用 theta 直接求解 ---
        elif s_curr < -EP and theta_curr is not None:
            # 尾方向: theta递增 (向外); 解析初值触发fsolve快速路径
            guess_in = theta_curr + target / (b * np.sqrt(theta_curr ** 2 + 1))
            theta_next = solve_next_handle_theta(
                theta_curr, target, b, direction=1,
                theta_guess=guess_in
            )
            if theta_next is None:
                return None
            arc = spiral_arc_length(theta_entry, theta_next, b)
            s_curr = -arc
            positions[i + 1] = spiral_point(theta_next, b)
            theta_curr = theta_next

        # --- 情形3: 在S曲线上或过渡区: 使用路径求解器 ---
        else:
            s_next = solve_next_handle_on_path(
                s_curr, target, path_obj, direction=-1
            )
            if s_next is None:
                return None
            P_next = get_unified_path_point(s_next, path_obj)
            if P_next is None:
                return None
            positions[i + 1] = P_next
            s_curr = s_next

            # 检测是否进入螺旋线段, 初始化theta跟踪
            if s_curr > s_total + EP:
                s_out = s_curr - s_total
                theta_curr = solve_theta_from_arc_length(
                    theta_entry, s_out, b, direction=1
                )
            elif s_curr < -EP:
                s_out = -s_curr
                theta_curr = solve_theta_from_arc_length(
                    theta_entry, s_out, b, direction=1
                )
            else:
                theta_curr = None

    return positions


# ============================================================
# 第2部分: 问题1 - 基础盘入模拟 (0~300s)
# ============================================================

def solve_problem1(output_dir='.'):
    """
    问题1: 沿螺距55cm的等距螺线顺时针盘入
    初始: 龙头前把手在螺线第16圈A点 (theta = 32π)
    速度: 1 m/s
    时间: 0~300s, 每秒输出
    """
    print("=" * 60)
    print("问题1: 基础盘入模拟 (0~300s)")
    print("=" * 60)

    pitch = 0.55
    b = get_b(pitch)
    theta_initial = 32 * np.pi

    r_initial = b * theta_initial
    print(f"螺距: {pitch} m, b = {b:.8f}")
    print(f"初始角度: theta = {theta_initial:.4f} rad")
    print(f"初始半径: r = {r_initial:.4f} m")

    # DragonSimulation 使用 head_direction=-1 (盘入), chain_direction=+1 (向后递增θ)
    sim = DragonSimulation(b, theta_initial, head_direction=-1)

    times = np.arange(0, 301, 1)
    print(f"\n开始模拟，共 {len(times)} 个时间步...")

    start = time.time()
    thetas_all, positions_all, times_out = sim.forward_simulation(
        times, head_speed=1.0, verbose=True
    )
    elapsed = time.time() - start
    n_steps = len(times_out)
    print(f"\n模拟完成，耗时 {elapsed:.2f}s，完成 {n_steps} 步")

    if n_steps < 2:
        print("错误: 模拟未能完成")
        return None

    # 计算速度
    speeds_all = DragonSimulation.compute_velocities(positions_all, times_out)

    # --- 生成 result1.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "位置和速度"

    # 表头
    headers = ["时间(s)"]
    for idx in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(
            idx, f"第{idx}节(前)")
        headers.append(f"{label}_x")
        headers.append(f"{label}_y")
    headers.append("时间(s)")
    for idx in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(
            idx, f"第{idx}节(前)")
        headers.append(label)

    ws.append(headers)

    # 使用np.round覆盖量化, 避免逐元素Python round()
    pos_rounded = np.round(positions_all, 6)
    spd_rounded = np.round(speeds_all, 6)
    for t_idx in range(n_steps):
        row = [times_out[t_idx]]
        # 展平 224x2 → 448个值
        row.extend(pos_rounded[t_idx].ravel().tolist())
        row.append(times_out[t_idx])
        row.extend(spd_rounded[t_idx].tolist())
        ws.append(row)

    result_path = os.path.join(output_dir, 'result1.xlsx')
    wb.save(result_path)
    print(f"结果已保存到: {result_path}")

    # --- 论文输出 ---
    print("\n" + "-" * 40)
    print("论文结果输出")
    print("-" * 40)

    report_times = [0, 60, 120, 180, 240, 300]
    report_handles = [0, 1, 51, 101, 151, 201, 223]

    valid_times = [t for t in report_times if t in times_out]

    print("\n表1: 位置 (m)")
    header_line = f"{'时间':>8s}"
    for h in report_handles:
        label = {0: "龙头", 223: "龙尾(后)"}.get(h, f"第{h}节")
        header_line += f"  {label}_x  {label}_y"
    print(header_line)

    for t in valid_times:
        t_idx = np.where(times_out == t)[0][0]
        line = f"{t:8.0f}s"
        for h in report_handles:
            line += f"  {positions_all[t_idx, h, 0]:8.4f}  {positions_all[t_idx, h, 1]:8.4f}"
        print(line)

    print("\n表2: 速度 (m/s)")
    header_line = f"{'时间':>8s}"
    for h in report_handles:
        label = {0: "龙头", 223: "龙尾(后)"}.get(h, f"第{h}节")
        header_line += f"  {label:>10s}"
    print(header_line)

    for t in valid_times:
        t_idx = np.where(times_out == t)[0][0]
        line = f"{t:8.0f}s"
        for h in report_handles:
            line += f"  {speeds_all[t_idx, h]:10.6f}"
        print(line)

    return thetas_all, positions_all, speeds_all, times_out, sim


# ============================================================
# 第3部分: 问题2 - 碰撞检测与终止时刻
# ============================================================

def check_collision_bench_polygons(thetas, b, bench_width=0.30):
    """
    使用多边形近似检查板凳碰撞
    每节板凳用4个顶点表示矩形
    返回: (collided: bool, idx1: int, idx2: int)
    """
    positions = spiral_points(thetas, b)
    half_w = bench_width / 2

    # 构建每个板凳的矩形顶点
    bench_verts = []
    for i in range(N_HANDLES - 1):
        p1, p2 = positions[i], positions[i + 1]
        v = p2 - p1
        len_v = np.linalg.norm(v)
        if len_v < 1e-10:
            rect = np.array([
                p1 + np.array([-half_w, -half_w]),
                p1 + np.array([half_w, -half_w]),
                p1 + np.array([half_w, half_w]),
                p1 + np.array([-half_w, half_w]),
            ])
        else:
            u = v / len_v
            n = np.array([-u[1], u[0]])
            nw = n * half_w
            rect = np.array([
                p1 - nw, p2 - nw, p2 + nw, p1 + nw,
            ])
        bench_verts.append(rect)

    # 检查非直接相邻的板凳对 (跳过 i+1, i+2 即跳过共用接头)
    # 注意: 当螺距很小时，不同圈之间的板凳可能碰撞 (链索引差 > 15)
    # 因此需要检查所有非相邻的板凳对
    for i in range(len(bench_verts) - 2):
        for j in range(i + 2, len(bench_verts)):
            if _polygons_intersect(bench_verts[i], bench_verts[j]):
                return True, i, j

    return False, -1, -1


def _polygons_intersect(poly1, poly2):
    """检查两个凸多边形是否相交 (含边交叉检测 + AABB预检)"""
    # 快速AABB预检: 如果外接矩形不相交则直接返回
    min1 = np.min(poly1, axis=0)
    max1 = np.max(poly1, axis=0)
    min2 = np.min(poly2, axis=0)
    max2 = np.max(poly2, axis=0)
    if (max1[0] < min2[0] or max2[0] < min1[0] or
        max1[1] < min2[1] or max2[1] < min1[1]):
        return False

    # 方法1: 顶点包含检测 (一个多边形的顶点在另一个内部)
    for pt in poly1:
        if _point_in_convex_polygon(pt, poly2):
            return True
    for pt in poly2:
        if _point_in_convex_polygon(pt, poly1):
            return True

    # 方法2: 边交叉检测 (捕获十字交叉等无顶点包含的情况)
    n1, n2 = len(poly1), len(poly2)
    for i in range(n1):
        a, b = poly1[i], poly1[(i + 1) % n1]
        for j in range(n2):
            c, d = poly2[j], poly2[(j + 1) % n2]
            if _segments_intersect(a, b, c, d):
                return True

    return False


def _segments_intersect(p1, p2, p3, p4):
    """检查线段 (p1-p2) 和 (p3-p4) 是否相交 (含端点接触)"""
    d1 = (p2[0] - p1[0]) * (p3[1] - p1[1]) - (p2[1] - p1[1]) * (p3[0] - p1[0])
    d2 = (p2[0] - p1[0]) * (p4[1] - p1[1]) - (p2[1] - p1[1]) * (p4[0] - p1[0])
    d3 = (p4[0] - p3[0]) * (p1[1] - p3[1]) - (p4[1] - p3[1]) * (p1[0] - p3[0])
    d4 = (p4[0] - p3[0]) * (p2[1] - p3[1]) - (p4[1] - p3[1]) * (p2[0] - p3[0])

    # 一般情形: 跨立测试
    if d1 * d2 < 0 and d3 * d4 < 0:
        return True

    # 端点共线且在线段上的情形
    eps = 1e-14
    if abs(d1) < eps and _point_on_segment(p3, p1, p2):
        return True
    if abs(d2) < eps and _point_on_segment(p4, p1, p2):
        return True
    if abs(d3) < eps and _point_on_segment(p1, p3, p4):
        return True
    if abs(d4) < eps and _point_on_segment(p2, p3, p4):
        return True

    return False


def _point_on_segment(pt, a, b):
    """检查点pt是否在线段ab上 (假设已共线)"""
    return (min(a[0], b[0]) - 1e-14 <= pt[0] <= max(a[0], b[0]) + 1e-14 and
            min(a[1], b[1]) - 1e-14 <= pt[1] <= max(a[1], b[1]) + 1e-14)


def _point_in_convex_polygon(pt, polygon):
    """
    检查点是否在凸多边形内部 (含边界)
    自动适应 CW/CCW 顶点顺序
    """
    n = len(polygon)
    signs = []
    for i in range(n):
        a = polygon[i]
        b = polygon[(i + 1) % n]
        cross = (b[0] - a[0]) * (pt[1] - a[1]) - (b[1] - a[1]) * (pt[0] - a[0])
        signs.append(cross)

    # 点在内部当且仅当所有有向面积同号 (都在边界同侧)
    # CCW时 cross >= 0, CW时 cross <= 0
    positive = all(s >= -1e-14 for s in signs)
    negative = all(s <= 1e-14 for s in signs)
    return positive or negative


def solve_problem2(thetas_all, positions_all, speeds_all, times_out, sim,
                   output_dir='.'):
    """
    问题2: 寻找舞龙队盘入的终止时刻 (碰撞发生前)
    从问题1结果继续模拟
    """
    print("\n" + "=" * 60)
    print("问题2: 碰撞检测与终止时刻")
    print("=" * 60)

    b = sim.b
    start_time_val = times_out[-1]

    # 继续模拟
    all_thetas = list(thetas_all)
    all_positions = list(positions_all)
    all_speeds = list(speeds_all)
    all_times = list(times_out)

    collision_time = None
    max_steps = 500

    print(f"从 T={start_time_val:.1f}s 继续模拟...")

    prev_thetas = all_thetas[-1] if len(all_thetas) > 0 else None

    for step in range(max_steps):
        current_time = start_time_val + (step + 1)
        if current_time > 1000:
            break

        # 龙头弧长 = 时间 (1 m/s), 使用hot-start加速
        result = sim.simulate_at_time(current_time, head_speed=1.0,
                                      prev_thetas=prev_thetas)
        if result is None:
            print(f"  T={current_time:.1f}s: 链式约束求解失败")
            collision_time = current_time
            break

        new_thetas, new_positions = result
        prev_thetas = new_thetas

        # 碰撞检测
        collided, idx1, idx2 = check_collision_bench_polygons(new_thetas, b)
        if collided:
            print(f"  T={current_time:.1f}s: 发生碰撞 (板凳 {idx1} vs {idx2})")
            collision_time = current_time
            break

        all_thetas.append(new_thetas)
        all_positions.append(new_positions)
        all_times.append(current_time)

        if step % 20 == 0:
            r_head = np.linalg.norm(new_positions[0])
            r_tail = np.linalg.norm(new_positions[-1])
            print(f"  T={current_time:.1f}s: r_head={r_head:.4f}m, r_tail={r_tail:.4f}m")

    # 速度计算
    n_steps = len(all_times)
    all_positions_arr = np.array(all_positions)
    all_speeds_arr = DragonSimulation.compute_velocities(all_positions_arr, np.array(all_times))

    if collision_time is not None:
        final_idx = max(0, n_steps - 2)  # 碰撞前一步
        print(f"\n碰撞时刻: T = {collision_time:.1f} s")
    else:
        final_idx = -1
        print(f"\n模拟{max_steps}步后未发生碰撞")

    final_thetas = all_thetas[final_idx]
    final_positions = all_positions[final_idx]
    final_speeds = all_speeds_arr[final_idx]
    final_time = all_times[final_idx]

    print(f"终止时刻: T = {final_time:.1f} s")
    print(f"龙头半径: r = {np.linalg.norm(final_positions[0]):.4f} m")
    print(f"龙尾(后)半径: r = {np.linalg.norm(final_positions[-1]):.4f} m")

    # --- 生成 result2.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "终止时刻位置和速度"
    ws.append(["终止时刻位置与速度"])
    ws.append([f"终止时刻: T = {final_time:.1f}s"])
    ws.append([])

    ws.append(["位置 (m)"])
    pos_headers = ["把手"]
    for h in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(h, f"第{h}节(前)")
        pos_headers.append(label)
    ws.append(pos_headers)

    x_row = ["x"] + [round(final_positions[h, 0], 6) for h in range(N_HANDLES)]
    y_row = ["y"] + [round(final_positions[h, 1], 6) for h in range(N_HANDLES)]
    ws.append(x_row)
    ws.append(y_row)

    ws.append([])
    ws.append(["速度 (m/s)"])
    vel_headers = ["把手"]
    for h in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(h, f"第{h}节(前)")
        vel_headers.append(label)
    ws.append(vel_headers)

    speed_row = ["速度"] + [round(final_speeds[h], 6) for h in range(N_HANDLES)]
    ws.append(speed_row)

    result_path = os.path.join(output_dir, 'result2.xlsx')
    wb.save(result_path)
    print(f"结果已保存到: {result_path}")

    # --- 论文输出 ---
    print("\n" + "-" * 40)
    print(f"终止时刻 T = {final_time:.1f}s 的关键输出")
    print("-" * 40)

    report_handles = [0, 1, 51, 101, 151, 201, 223]
    print(f"\n位置 (m):")
    for h in report_handles:
        label = {0: "龙头(前)", 223: "龙尾(后)"}.get(h, f"第{h}节(身)")
        print(f"  {label}: ({final_positions[h, 0]:.6f}, {final_positions[h, 1]:.6f})")

    print(f"\n速度 (m/s):")
    for h in report_handles:
        label = {0: "龙头(前)", 223: "龙尾(后)"}.get(h, f"第{h}节(身)")
        print(f"  {label}: {final_speeds[h]:.6f}")

    return collision_time, final_thetas, final_positions, final_speeds, final_time


# ============================================================
# 第4部分: 问题3 - 最小螺距确定 (使用修正后的链方向)
# ============================================================

def solve_problem3(output_dir='.'):
    """
    问题3: 确定最小螺距，使龙头能盘入到调头空间边界 (r = 4.5m)

    修正: 使用 DragonSimulation 的链方向修正
    """
    print("\n" + "=" * 60)
    print("问题3: 最小螺距确定")
    print("=" * 60)

    turning_radius = 4.5
    theta_initial = 32 * np.pi

    print(f"调头空间半径: {turning_radius} m")

    def can_reach_boundary(pitch):
        """判断龙头能否沿螺线盘入到 r = 4.5m 边界"""
        b = get_b(pitch)
        theta_boundary = turning_radius / b

        if theta_boundary > theta_initial:
            return False

        # 使用修正后的 DragonSimulation: head_direction=-1 (盘入)
        sim = DragonSimulation(b, theta_initial, head_direction=-1)
        thetas = sim.compute_chain_positions(theta_boundary)

        if thetas is None:
            return False
        if np.any(thetas < 0):
            return False
        if check_collision_bench_polygons(thetas, b)[0]:
            return False

        return True

    # 二分搜索
    pitch_low, pitch_high = 0.1, 5.0
    print(f"搜索范围: 螺距 [{pitch_low}, {pitch_high}] m\n")

    iteration = 0
    while pitch_high - pitch_low > 0.001 and iteration < 50:
        pitch_mid = (pitch_low + pitch_high) / 2
        if can_reach_boundary(pitch_mid):
            pitch_high = pitch_mid
            print(f"  迭代 {iteration + 1}: p = {pitch_mid:.4f}m → 可达 ✓")
        else:
            pitch_low = pitch_mid
            print(f"  迭代 {iteration + 1}: p = {pitch_mid:.4f}m → 不可达 ✗")
        iteration += 1

    min_pitch = pitch_high
    print(f"\n最小螺距: {min_pitch:.6f} m = {min_pitch*100:.4f} cm")

    # 最终状态
    b_opt = get_b(min_pitch)
    theta_boundary = turning_radius / b_opt

    sim_opt = DragonSimulation(b_opt, theta_initial, head_direction=-1)
    final_thetas = sim_opt.compute_chain_positions(theta_boundary)
    final_positions = spiral_points(final_thetas, b_opt)

    print(f"\n龙头到达边界时的状态:")
    print(f"  边界角度: {theta_boundary:.4f} rad")
    print(f"  边界半径: {b_opt * theta_boundary:.4f} m")
    print(f"  龙尾(后)半径: {b_opt * final_thetas[-1]:.4f} m")

    # 保存 result3.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "最小螺距结果"
    ws.append(["问题3: 最小螺距"])
    ws.append([f"最小螺距: {min_pitch:.6f} m"])
    ws.append([f"调头空间半径: {turning_radius} m"])
    ws.append([])
    ws.append(["龙头到达边界时的龙队状态"])
    ws.append(["把手", "角度(rad)", "半径(m)", "x(m)", "y(m)"])

    for h in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(h, f"第{h}节(前)")
        ws.append([
            label,
            round(final_thetas[h], 6),
            round(b_opt * final_thetas[h], 6),
            round(final_positions[h, 0], 6),
            round(final_positions[h, 1], 6)
        ])

    result_path = os.path.join(output_dir, 'result3.xlsx')
    wb.save(result_path)
    print(f"结果已保存到: {result_path}")

    return min_pitch, final_thetas, final_positions


# ============================================================
# 第5部分: 问题4 - S形调头曲线与完整模拟 (完全重写)
# ============================================================

def solve_problem4(output_dir='.'):
    """
    问题4: S形调头曲线设计 + 完整模拟

    关键改进:
    - 使用 CompositePath 统一管理三段路径 (盘入螺线 + S曲线 + 盘出螺线)
    - 使用弧长参数化方法正确求解所有把手的链约束
    - 龙身各把手通过路径弧长偏移确定位置

    盘入螺距: 1.7 m
    盘出螺线: 与盘入螺线关于螺线中心中心对称
    调头空间: 直径9m圆盘 (半径4.5m)
    """
    print("\n" + "=" * 60)
    print("问题4: S形调头曲线与完整模拟")
    print("=" * 60)

    pitch = 1.7
    R_turn = 4.5
    b = get_b(pitch)
    theta_initial = 32 * np.pi

    print(f"螺距: {pitch} m, b = {b:.6f}")
    print(f"调头空间半径: {R_turn} m")

    # 构建组合路径
    path = CompositePath(pitch, R_turn)

    print(f"\n盘入到达边界: theta={path.theta_entry:.4f} rad")
    print(f"  入口位置: ({path.P_entry[0]:.4f}, {path.P_entry[1]:.4f}) m")
    print(f"盘出起点: theta出口与入口相同, 中心对称像")
    print(f"  出口位置: ({path.P_exit[0]:.4f}, {path.P_exit[1]:.4f}) m")

    print(f"\nS曲线参数:")
    print(f"  R1 = {path.R1:.4f} m (前段), R2 = {path.R2:.4f} m (后段)")
    print(f"  Arc1长 = {path.arc1_len:.4f} m, Arc2长 = {path.arc2_len:.4f} m")
    print(f"  S曲线总长 = {path.s_total:.4f} m")
    print(f"  O1 = ({path.O1[0]:.4f}, {path.O1[1]:.4f})")
    print(f"  O2 = ({path.O2[0]:.4f}, {path.O2[1]:.4f})")
    print(f"  P_mid = ({path.P_mid[0]:.4f}, {path.P_mid[1]:.4f})")

    # --- 不同半径比的调头曲线长度分析 ---
    print("\n" + "-" * 40)
    print("调头曲线长度分析 (不同半径比)")
    print("-" * 40)

    ratios = [1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0, 8.0]
    valid_results = []

    for ratio in ratios:
        try:
            # 修改半径比重新求解
            def eq(R2):
                R1 = ratio * R2
                O1_t = path.P_entry + R1 * path.n_entry
                O2_t = path.P_exit + R2 * path.n_exit
                return np.linalg.norm(O2_t - O1_t) - (R1 + R2)

            Rl, Rh = 0.01, R_turn
            for _ in range(100):
                if eq(Rl) * eq(Rh) <= 0:
                    break
                Rh *= 0.5
            if eq(Rl) * eq(Rh) > 0:
                Rh = R_turn
                for _ in range(100):
                    Rh *= 1.5
                    if eq(Rl) * eq(Rh) <= 0:
                        break
                if eq(Rl) * eq(Rh) > 0:
                    print(f"  R1/R2 = {ratio:.1f}: 无法求解 ✗")
                    continue

            R2_t = brentq(eq, Rl, Rh, xtol=1e-10)
            R1_t = ratio * R2_t

            # 验证在边界内
            O1_t = path.P_entry + R1_t * path.n_entry
            O2_t = path.P_exit + R2_t * path.n_exit
            dir_t = (O2_t - O1_t) / np.linalg.norm(O2_t - O1_t)
            P_mid_t = O1_t + R1_t * dir_t

            if (np.linalg.norm(P_mid_t) > R_turn * 1.02 or
                np.linalg.norm(O1_t) > R_turn * 1.02):
                print(f"  R1/R2 = {ratio:.1f}: 超出调头空间边界 ✗")
                continue

            # 计算S曲线总长
            v1 = path.P_entry - O1_t
            a1 = np.arctan2(v1[1], v1[0])
            v_m = P_mid_t - O1_t
            a_m = np.arctan2(v_m[1], v_m[0])
            diff1 = path._angle_diff(a1, a_m)
            len1 = abs(diff1) * R1_t

            v_m2 = P_mid_t - O2_t
            a_m2 = np.arctan2(v_m2[1], v_m2[0])
            v_e = path.P_exit - O2_t
            a_e = np.arctan2(v_e[1], v_e[0])
            diff2 = path._angle_diff(a_m2, a_e)
            len2 = abs(diff2) * R2_t

            total = len1 + len2
            valid_results.append((ratio, total, R1_t, R2_t))
            status = "✓" if np.linalg.norm(P_mid_t) <= R_turn else "△"
            print(f"  R1/R2 = {ratio:.1f}: R1={R1_t:.2f}, R2={R2_t:.2f}, "
                  f"S总长={total:.3f}m {status}")

        except Exception as e:
            print(f"  R1/R2 = {ratio:.1f}: 出错 - {str(e)[:30]}")

    if valid_results:
        best = min(valid_results, key=lambda x: x[1])
        print(f"\n结论: R1:R2 = {best[0]:.1f} 时最短 ({best[1]:.3f}m)")
    else:
        print("\n结论: 无有效解")

    # --- 完整模拟: -100s 到 100s ---
    print("\n" + "-" * 40)
    print("进行完整模拟 (-100s ~ 100s)")
    print("-" * 40)

    all_times_4 = []
    all_positions_4 = []

    # 阶段1: t < 0, 沿盘入螺线运动 (使用 DragonSimulation 保证正确弦长约束)
    # t=0 时龙头在 P_entry, t<0 时龙头在 P_entry 外侧
    # 计算龙头在 t=-100s 时的位置 (从 P_entry 向外走100m)
    print("计算盘入阶段 (t < 0)...")

    # 头在最外侧时刻 (t=-100) 的角度:
    theta_far = solve_theta_from_arc_length(
        path.theta_entry, 100.0, b, direction=1
    )
    if theta_far is None or theta_far > theta_initial:
        theta_far = theta_initial

    # 创建从最外侧向内盘入的模拟器
    sim_entry = DragonSimulation(b, theta_far, head_direction=-1)

    # t从 -100 到 0 (模拟器内部时间: 0 到 100)
    neg_times = list(range(-100, 0))
    prev_thetas_entry = None

    for t in neg_times:
        sim_t = t + 100  # 映射到模拟器内部时间: 0→-100, 100→0
        result = sim_entry.simulate_at_time(sim_t, head_speed=1.0,
                                            prev_thetas=prev_thetas_entry)
        if result is None:
            print(f"  T={t}s: 盘入模拟失败")
            break

        thetas, positions = result
        prev_thetas_entry = thetas  # hot-start
        all_times_4.append(t)
        all_positions_4.append(positions)

        if t % 20 == 0:
            r_head = np.linalg.norm(positions[0])
            print(f"  T={t:.0f}s: r_head={r_head:.3f}m, "
                  f"theta={thetas[0]:.2f}rad")

    # t=0: S曲线入口 (使用弦长约束求解器保证精确)
    print("计算调头及盘出阶段 (t >= 0)...")

    pos_t0 = compute_chain_on_unified_path(0.0, path)
    if pos_t0 is not None:
        all_times_4.append(0)
        all_positions_4.append(pos_t0)
        print(f"  T=0s: 到达S曲线入口, "
              f"r_head={np.linalg.norm(pos_t0[0]):.4f}m")

    # t > 0: 通过S曲线然后盘出
    # 在统一路径上使用弦长约束求解器
    s_total = path.s_total
    print(f"  S曲线弧长: {s_total:.4f}m, 通过时间: {s_total/1.0:.2f}s")

    pos_times = range(1, 101)

    for t in pos_times:
        positions = compute_chain_on_unified_path(t, path)
        if positions is None:
            print(f"  T={t}s: 弦长约束求解失败, 跳过")
            continue

        all_times_4.append(t)
        all_positions_4.append(positions)

        if t % 20 == 0:
            r_head = np.linalg.norm(positions[0])
            r_tail = np.linalg.norm(positions[-1])
            phase = "S曲线" if t <= s_total else "盘出螺线"
            print(f"  T={t:.0f}s: r_head={r_head:.4f}m, "
                  f"r_tail={r_tail:.4f}m [{phase}]")

    # 转换为numpy数组
    all_times_4_arr = np.array(all_times_4)
    all_positions_4_arr = np.array(all_positions_4)
    n_steps_4 = len(all_times_4)

    # 计算速度
    all_speeds_4_arr = DragonSimulation.compute_velocities(
        all_positions_4_arr, all_times_4_arr
    )

    print(f"\n模拟完成: {n_steps_4} 个时间步 (T={all_times_4_arr[0]}s ~ {all_times_4_arr[-1]}s)")

    # 保存 result4.xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "位置和速度"

    headers = ["时间(s)"]
    for idx in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(
            idx, f"第{idx}节(前)")
        headers.append(f"{label}_x")
        headers.append(f"{label}_y")
    headers.append("时间(s)")
    for idx in range(N_HANDLES):
        label = {0: "龙头(前)", N_HANDLES - 1: "龙尾(后)"}.get(
            idx, f"第{idx}节(前)")
        headers.append(label)

    ws.append(headers)

    pos4_rounded = np.round(all_positions_4_arr, 6)
    spd4_rounded = np.round(all_speeds_4_arr, 6)
    for t_idx in range(n_steps_4):
        row = [all_times_4_arr[t_idx]]
        row.extend(pos4_rounded[t_idx].ravel().tolist())
        row.append(all_times_4_arr[t_idx])
        row.extend(spd4_rounded[t_idx].tolist())
        ws.append(row)

    result_path = os.path.join(output_dir, 'result4.xlsx')
    wb.save(result_path)
    print(f"结果已保存到: {result_path}")

    # --- 论文输出 ---
    print("\n" + "-" * 40)
    print("论文结果输出 (问题4)")
    print("-" * 40)

    report_times_4 = [-100, -50, 0, 50, 100]
    report_handles = [0, 1, 51, 101, 151, 201, 223]

    valid_times_4 = [t for t in report_times_4 if t in all_times_4_arr]

    print("\n位置 (m):")
    for t in valid_times_4:
        t_idx = np.where(all_times_4_arr == t)[0][0]
        print(f"  T={t:4d}s:", end="")
        for h in report_handles:
            label = {0: "龙头", 223: "龙尾(后)"}.get(h, f"身{h}")
            print(f" {label}=({all_positions_4_arr[t_idx, h, 0]:.4f},{all_positions_4_arr[t_idx, h, 1]:.4f})", end="")
        print()

    print("\n速度 (m/s):")
    for t in valid_times_4:
        t_idx = np.where(all_times_4_arr == t)[0][0]
        print(f"  T={t:4d}s:", end="")
        for h in report_handles:
            label = {0: "龙头", 223: "龙尾(后)"}.get(h, f"身{h}")
            print(f" {label}={all_speeds_4_arr[t_idx, h]:.6f}", end="")
        print()

    return (all_times_4_arr, all_positions_4_arr, all_speeds_4_arr,
            path)


# ============================================================
# 第6部分: 问题5 - 最大龙头速度
# ============================================================

def solve_problem5(all_times_4, all_positions_4, all_speeds_4, path,
                   output_dir='.'):
    """
    问题5: 确定龙头的最大行进速度

    约束: 所有把手速度 <= 2 m/s
    龙头速度 v_head 恒定 (沿问题4路径)
    通过分析链上各把手的速度放大因子确定最大允许龙头速度
    """
    print("\n" + "=" * 60)
    print("问题5: 最大龙头行进速度确定")
    print("=" * 60)

    # 问题4的模拟中速度=1m/s, 速度数组 = 速度放大因子
    n_steps = len(all_times_4)

    # 计算每个步的所有把手速度与龙头速度的比值
    speed_ratios = np.zeros_like(all_speeds_4)
    for t_idx in range(n_steps):
        head_speed = all_speeds_4[t_idx, 0]
        if head_speed > 1e-10:
            speed_ratios[t_idx] = all_speeds_4[t_idx] / head_speed
        else:
            speed_ratios[t_idx] = 1.0

    # 找到整个过程中最大的速度放大因子
    # 只考虑非头的把手 (索引1+)
    max_ratio_overall = np.max(speed_ratios[:, 1:])

    # 找到最大速度发生的时间步和把手
    max_idx = np.unravel_index(np.argmax(speed_ratios[:, 1:], axis=None),
                                speed_ratios[:, 1:].shape)
    t_max = all_times_4[max_idx[0]]
    handle_max = max_idx[1] + 1  # +1 because we excluded head

    print(f"\n速度放大因子分析 (v_head=1m/s时):")
    print(f"  最大放大因子: {max_ratio_overall:.6f}x")
    print(f"  发生时刻: T={t_max:.0f}s")
    print(f"  发生位置: 第{handle_max}节把手")

    # 最大龙头速度
    max_head_speed = 2.0 / max_ratio_overall
    print(f"\n龙头最大速度: {max_head_speed:.6f} m/s")

    # 分析S曲线段的放大因子
    s_total = path.s_total
    print(f"\nS曲线段 (弧长 {s_total:.2f}m) 速度分析:")
    for t_idx in range(n_steps):
        t = all_times_4[t_idx]
        if 0 <= t <= s_total + 5:
            max_rat = np.max(speed_ratios[t_idx, 1:])
            if max_rat > 1.5:  # 仅打印放大显著的
                print(f"  T={t:.0f}s: 放大因子 = {max_rat:.4f}x")

    # 保存结果
    print(f"\n结论: 龙头最大行进速度 = {max_head_speed:.6f} m/s")

    # 保存速度分析数据
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "速度分析"
    ws.append(["速度分析"])
    ws.append([f"龙头最大速度: {max_head_speed:.6f} m/s"])
    ws.append([f"最大放大因子: {max_ratio_overall:.6f}x"])
    ws.append([f"发生时刻: T={t_max}s, 第{handle_max}节把手"])
    ws.append([])
    ws.append(["时间(s)", "龙头速度(m/s)", "最大速度(m/s)", "最大放大因子"])
    for t_idx in range(n_steps):
        t = all_times_4[t_idx]
        head_sp = all_speeds_4[t_idx, 0]
        max_sp = np.max(all_speeds_4[t_idx])
        max_rat = np.max(speed_ratios[t_idx, 1:])
        ws.append([round(t, 1), round(head_sp, 6), round(max_sp, 6),
                   round(max_rat, 6)])

    result_path = os.path.join(output_dir, 'result5_speed_analysis.xlsx')
    wb.save(result_path)
    print(f"速度分析已保存到: {result_path}")

    return max_head_speed, max_ratio_overall


# ============================================================
# 主函数
# ============================================================

def main():
    print("=" * 70)
    print("2024年高教社杯全国大学生数学建模竞赛 A题 '板凳龙'闹元宵")
    print("完整求解程序 (v2.4 - 优化版)")
    print("=" * 70)

    output_dir = os.path.dirname(os.path.abspath(__file__))
    if not output_dir:
        output_dir = '.'
    print(f"\n输出目录: {output_dir}")

    plt.rcParams['font.sans-serif'] = ['SimHei', 'DejaVu Sans']
    plt.rcParams['axes.unicode_minus'] = False

    # ========================
    # 问题1
    # ========================
    print("\n" + ">" * 30 + " 问题1 " + "<" * 30)
    try:
        results1 = solve_problem1(output_dir)
        if results1 is not None:
            thetas_all, positions_all, speeds_all, times_out, sim = results1
        else:
            thetas_all = positions_all = speeds_all = times_out = sim = None
    except Exception as e:
        print(f"问题1 出错: {e}")
        import traceback
        traceback.print_exc()
        thetas_all = positions_all = speeds_all = times_out = sim = None

    # ========================
    # 问题2
    # ========================
    print("\n" + ">" * 30 + " 问题2 " + "<" * 30)
    if thetas_all is not None and len(thetas_all) > 0:
        try:
            solve_problem2(thetas_all, positions_all, speeds_all,
                          times_out, sim, output_dir)
        except Exception as e:
            print(f"问题2 出错: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("问题2 跳过: 缺乏问题1的模拟结果")

    # ========================
    # 问题3
    # ========================
    print("\n" + ">" * 30 + " 问题3 " + "<" * 30)
    try:
        solve_problem3(output_dir)
    except Exception as e:
        print(f"问题3 出错: {e}")
        import traceback
        traceback.print_exc()

    # ========================
    # 问题4
    # ========================
    print("\n" + ">" * 30 + " 问题4 " + "<" * 30)
    try:
        results4 = solve_problem4(output_dir)
        if results4 is not None:
            (all_times_4, all_positions_4, all_speeds_4, path) = results4
        else:
            all_times_4 = all_positions_4 = all_speeds_4 = path = None
    except Exception as e:
        print(f"问题4 出错: {e}")
        import traceback
        traceback.print_exc()
        all_times_4 = all_positions_4 = all_speeds_4 = path = None

    # ========================
    # 问题5
    # ========================
    print("\n" + ">" * 30 + " 问题5 " + "<" * 30)
    if all_positions_4 is not None and path is not None:
        try:
            solve_problem5(all_times_4, all_positions_4, all_speeds_4,
                          path, output_dir)
        except Exception as e:
            print(f"问题5 出错: {e}")
            import traceback
            traceback.print_exc()
    else:
        print("问题5 跳过: 缺乏问题4的结果")

    # ========================
    # 绘图
    # ========================
    print("\n" + ">" * 30 + " 绘图 " + "<" * 30)
    try:
        fig, axes = plt.subplots(2, 2, figsize=(14, 12))

        # 问题1轨迹
        if thetas_all is not None:
            ax = axes[0, 0]
            ax.set_title('问题1: 龙队盘入轨迹 (0~300s)')
            for step in range(0, len(times_out), 30):
                pts = positions_all[step]
                ax.plot(pts[:, 0], pts[:, 1], 'b-', alpha=0.3, lw=0.8)
            ax.plot(positions_all[0, :, 0], positions_all[0, :, 1],
                    'g-', lw=2, label='t=0s')
            ax.plot(positions_all[-1, :, 0], positions_all[-1, :, 1],
                    'r-', lw=2, label=f't={times_out[-1]:.0f}s')
            ax.axhline(0, color='gray', lw=0.5)
            ax.axvline(0, color='gray', lw=0.5)
            ax.set_aspect('equal')
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        # 问题4轨迹
        if all_positions_4 is not None:
            ax = axes[0, 1]
            ax.set_title('问题4: 调头模拟 (-100s~100s)')
            # 绘制调头空间
            theta_circle = np.linspace(0, 2*np.pi, 100)
            ax.plot(4.5*np.cos(theta_circle), 4.5*np.sin(theta_circle),
                    'k--', alpha=0.5, label='调头空间')
            # 绘制几个时刻的龙队
            for t_mark in [-100, -50, 0, 50, 100]:
                if t_mark in all_times_4:
                    t_idx = np.where(all_times_4 == t_mark)[0][0]
                    pts = all_positions_4[t_idx]
                    ax.plot(pts[:, 0], pts[:, 1], '-', lw=1.5,
                            label=f'T={t_mark}s')
            # 绘制S曲线
            s_samples = np.linspace(0, path.s_total, 200)
            s_pts = np.array([path.point_on_curve(s) for s in s_samples])
            ax.plot(s_pts[:, 0], s_pts[:, 1], 'm-', lw=3, alpha=0.7,
                    label='S曲线')
            ax.axhline(0, color='gray', lw=0.5)
            ax.axvline(0, color='gray', lw=0.5)
            ax.set_aspect('equal')
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        # 速度分析
        if speeds_all is not None:
            ax = axes[1, 0]
            ax.set_title('问题1: 龙头与龙尾速度')
            ax.plot(times_out, speeds_all[:, 0], 'r-', label='龙头')
            ax.plot(times_out, speeds_all[:, -1], 'b-', label='龙尾(后)')
            ax.set_xlabel('时间 (s)')
            ax.set_ylabel('速度 (m/s)')
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        # 问题4速度
        if all_speeds_4 is not None:
            ax = axes[1, 1]
            ax.set_title('问题4: 龙头与龙尾速度')
            ax.plot(all_times_4, all_speeds_4[:, 0], 'r-', label='龙头')
            ax.plot(all_times_4, all_speeds_4[:, -1], 'b-', label='龙尾(后)')
            ax.axvline(0, color='gray', ls='--', alpha=0.5)
            ax.set_xlabel('时间 (s)')
            ax.set_ylabel('速度 (m/s)')
            ax.legend(fontsize=8)
            ax.grid(True, alpha=0.3)

        plt.tight_layout()
        plt_path = os.path.join(output_dir, 'simulation_plots.png')
        plt.savefig(plt_path, dpi=150)
        print(f"轨迹图已保存到: {plt_path}")
    except Exception as e:
        print(f"绘图出错: {e}")
        import traceback
        traceback.print_exc()

    print("\n" + "=" * 70)
    print("所有问题求解完成")
    print("=" * 70)
    print("\n生成的结果文件:")
    print("  - result1.xlsx            (问题1: 0~300s盘入模拟)")
    print("  - result2.xlsx            (问题2: 终止时刻状态)")
    print("  - result3.xlsx            (问题3: 最小螺距)")
    print("  - result4.xlsx            (问题4: -100s~100s模拟)")
    print("  - result5_speed_analysis.xlsx (问题5: 速度分析)")
    print("  - simulation_plots.png    (轨迹可视化)")


if __name__ == '__main__':
    main()
